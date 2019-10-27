import openpyxl
from common.reques import Request
import json
from common import contants
import json
class Case:
    def __init__(self):
        '''定义一个类描述测试数据'''
        self.case_id = None
        self.title = None
        self.url = None
        self.method =None
        self.data = None
        self.expected = None

class DoExcel:
    def __init__(self,file_name):
        self.file_name = file_name
        '''初始化函数内打开Excel文件'''
        try:
            self.workbook = openpyxl.load_workbook(filename=file_name)
        except FileNotFoundError as e:
            print('文件不存在，请检查文件路径')
            raise e

    def get_cases(self,sheet_name):
        '''定位表单'''
        sheet = self.workbook[sheet_name]
        #获取最大的行数
        max_row = sheet.max_row
        cases =[]
        for i in range(2,max_row+1):
            case=Case()
            case.case_id = sheet.cell(row=i,column =1).value#取第i行，第一格的值
            case.title = sheet.cell(row=i, column=2).value
            case.url = sheet.cell(row=i, column=3).value
            case.data = sheet.cell(row=i, column=4).value
            case.method = sheet.cell(row=i, column=5).value
            case.expected = sheet.cell(row=i, column=6).value
            cases.append(case)
        return cases
    def get_sheet_name(self):
        '''获取到workbook 里面所有的sheet名称的列表'''
        return self.workbook.sheetnames

    def write_actual_case_id(self,sheet_name,case_id,actual):
        #根据sheet_name 定位sheet，然后根据case_id定位到行，取到当前actual这个单元格
        #然后赋值，在保存当前的workboook
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        for r in range(2,max_row+1):
            case_id_r = sheet.cell(r,1).value#获取第r行第1列，也就是获取caseid
            if case_id_r ==case_id:#判断excel内取到当前行的case_id是否等于传进来的case_id
                sheet.cell(r,7).value=actual#写入传进来的实际结果到当前行的actual列的单元格
                self.workbook.save(filename=self.file_name)
                break
    def write_result_case_id(self,sheet_name,case_id,actual):
        #根据sheet_name 定位sheet，然后根据case_id定位到行，取到当前result这个单元格
        #然后赋值，在保存当前的workboook
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        for r in range(2,max_row+1):
            case_id_r = sheet.cell(r,1).value
            if case_id_r ==case_id:
                sheet.cell(r,8).value=actual
                self.workbook.save(filename=self.file_name)
                break

if __name__ == '__main__':
    do_excel = DoExcel(contants.case_file)
    #获取表单里面所有的sheet名称的列表
    sheet_names = do_excel.get_sheet_name()
    print("sheet名称列表",sheet_names)
    case_list= ['login','register']#定义一个执行测试用例列表
    for sheet_name in sheet_names:
        if sheet_name in case_list:#如果当前sheet_name 不在可执行的case_list内，则不执行
            cases1 = do_excel.get_cases(sheet_name)#定位表单
            print(sheet_name+'测试用例个数',len(cases1))
            for case in cases1:
                print('case信息：',case.__dict__)#打印case信息
                data = eval(case.data)#Excel里面取到data是一个字符串，使用eval()函数将字符串转化成字典
                # data = json.loads(case.data,encoding="GBK")
                resp = Request(methon =case.method,url=case.url,data=data)#通过封装Request类来完成接口的调用
                print(resp.get_text())
                print(resp.get_status_code())#打印响应码
                resp_text = json.dumps(resp.get_json(),ensure_ascii=False,indent=4)#通过json。dumps函数将字典转成格式化后的字符串
                print(resp_text)
                do_excel.write_actual_case_id(sheet_name=sheet_name,case_id=case.case_id,actual=resp.get_text())#将响应的str类型的实际结果写入表格
                #判断接口响应是否和Excel里面的expected是否一致
                if case.expected == resp.get_text():
                    print('结果:','pass')
                    do_excel.write_result_case_id(sheet_name=sheet_name,case_id=case.case_id,actual='pass')
                else:
                    print('结果:',"Fail")
                    do_excel.write_result_case_id(sheet_name=sheet_name, case_id=case.case_id, actual='fail')